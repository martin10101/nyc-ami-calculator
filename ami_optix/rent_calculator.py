"""
Utility helpers for loading the annual AMI rent workbook and deriving
monthly/annual net rents based on utility selections.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Tuple, Any
import os
import math
import re

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

BEDROOM_LABELS = ["studio", "1 BR", "2 BR", "3 BR", "4 BR", "5 BR"]

# NYC HPD regulatory haircut at the 100% AMI band: max collectable rent is
# 97% of the published 100% AMI rent. Applied to gross rent at the source so
# every downstream consumer (solver rent objective, display, edge scenarios)
# uses the corrected number.
HAIRCUT_BAND_AMI = 1.0
HAIRCUT_FACTOR = 0.97

COOKING_OPTIONS = {
    "electric": "Electric Stove",
    "gas": "Gas Stove",
    "na": "N/A or owner pays",
}

HEAT_OPTIONS = {
    "electric_ccashp": "Electric Heat - Cold Climate Air Source Heat Pump (ccASHP)1",
    "electric_other": "Electric Heat - Other2",
    "gas": "Gas Heat",
    "oil": "Oil Heat",
    "na": "N/A or owner pays",
}

ELECTRICITY_OPTIONS = {
    "tenant_pays": "Tenant Pays",
    "na": "N/A or owner pays",
}

HOT_WATER_OPTIONS = {
    "electric_heat_pump": "Electric Hot Water - Heat Pump",
    "electric_other": "Electric Hot Water - Other",
    "gas": "Gas Hot Water",
    "oil": "Oil Hot Water",
    "na": "N/A or owner pays",
}


UTILITY_OPTION_MAP = {
    'electricity': ELECTRICITY_OPTIONS,
    'cooking': COOKING_OPTIONS,
    'heat': HEAT_OPTIONS,
    'hot_water': HOT_WATER_OPTIONS,
}

LABEL_TO_CATEGORY = {label: key for key, options in UTILITY_OPTION_MAP.items() for label in options.values()}
HEADER_TO_CATEGORY = {
    "apartment electricity only": "electricity",
    "cooking": "cooking",
    "heat": "heat",
    "hot water": "hot_water",
}

@dataclass
class RentSchedule:
    gross_rents: Dict[Tuple[float, str], float]
    allowances: Dict[str, Dict[str, Dict[str, float]]]

    def rent_for(self, ami_percent: float, bedrooms: float, selections: Dict[str, str]) -> float:
        return self.rent_components(ami_percent, bedrooms, selections)['net']

    def rent_components(self, ami_percent: float, bedrooms: float, selections: Dict[str, str]) -> Dict[str, Any]:
        bedroom_label = _normalize_bedroom_label(bedrooms)
        gross_pre_haircut = float(self._gross_rents_lookup(ami_percent, bedroom_label))
        # Apply the 100% AMI haircut at the source. At exactly 100% AMI the
        # landlord can collect at most 97% of the headline rent (NYC HPD rule).
        if abs(ami_percent - HAIRCUT_BAND_AMI) < 1e-6:
            gross = gross_pre_haircut * HAIRCUT_FACTOR
        else:
            gross = gross_pre_haircut
        allowances: Dict[str, Dict[str, Any]] = {}
        total_allowance = 0.0
        for category, options in UTILITY_OPTION_MAP.items():
            selection_key = (selections or {}).get(category, 'na')
            option_label = _resolve_option_label(category, selection_key)
            amount = float(self._allowance_lookup(category, option_label, bedroom_label))
            allowances[category] = {
                'category': category,
                'label': option_label,
                'amount': amount,
            }
            total_allowance += amount
        net = max(gross - total_allowance, 0.0)
        return {
            'gross': gross,
            'gross_pre_haircut': gross_pre_haircut,
            'allowances': allowances,
            'allowance_total': total_allowance,
            'net': net,
            'haircut_applied': gross != gross_pre_haircut,
        }

    def _gross_rents_lookup(self, ami_percent: float, bedroom_label: str) -> float:
        key = (round(ami_percent, 4), bedroom_label)
        if key not in self.gross_rents:
            raise ValueError(f"Rent table missing entry for {ami_percent*100:.0f}% AMI / {bedroom_label}.")
        return float(self.gross_rents[key])

    def _allowance_lookup(self, category: str, option: str, bedroom_label: str) -> float:
        cat = self.allowances.get(category)
        if not cat or option not in cat:
            return 0.0
        bedroom_allowances = cat[option]
        return float(bedroom_allowances.get(bedroom_label, 0.0))


def _pandas_engine_for(path: str) -> str | None:
    ext = os.path.splitext(path.lower())[1]
    if ext == ".xlsb":
        return "pyxlsb"
    return None


def load_excel_file(file_path: str):
    """Load an Excel workbook with macro preservation and faster parsing.

    Uses ``data_only=True`` to avoid parsing formulas and ``keep_links=False``
    to skip external link binding so large templates load noticeably faster.
    """
    if file_path.lower().endswith('.xlsm'):
        return load_workbook(
            file_path,
            keep_vba=True,
            data_only=True,
            keep_links=False,
        )
    return load_workbook(
        file_path,
        data_only=True,
        keep_links=False,
    )


_IF_FORMULA_RE = re.compile(r'=IF\s*\([^,]+,\s*([\d.]+)\s*,\s*0\s*\)')


def _extract_if_formula_values(workbook_path: str) -> Dict[Tuple[int, int], float]:
    """Extract true-branch values from ``=IF(..., VALUE, 0)`` formulas.

    Rent calculator workbooks use conditional formulas so that unselected
    utility options display 0.  When loaded with ``data_only=True`` the cached
    result is 0, making the server think the allowance is zero.  This function
    loads the workbook *without* ``data_only`` and pulls the actual numeric
    value out of each IF formula, keyed by ``(pandas_0row, pandas_0col)``.
    """
    # openpyxl cannot read .xlsb formulas.
    if workbook_path.lower().endswith('.xlsb'):
        return {}
    try:
        wb = load_workbook(workbook_path, data_only=False, keep_links=False)
        ws = wb['AMI & Rent']
    except Exception:
        return {}

    values: Dict[Tuple[int, int], float] = {}
    try:
        max_row = min(ws.max_row or 0, 30)
        max_col = min(ws.max_column or 0, 20)
        for row_1 in range(1, max_row + 1):
            for col_1 in range(1, max_col + 1):
                cell_val = ws.cell(row=row_1, column=col_1).value
                if isinstance(cell_val, str) and cell_val.startswith('='):
                    m = _IF_FORMULA_RE.search(cell_val)
                    if m:
                        values[(row_1 - 1, col_1 - 1)] = float(m.group(1))
    except Exception:
        pass
    finally:
        wb.close()
    return values


def load_rent_schedule(workbook_path: str) -> RentSchedule:
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(f"Rent calculator workbook not found: {workbook_path}")

    engine = _pandas_engine_for(workbook_path)
    sheet = pd.read_excel(workbook_path, sheet_name="AMI & Rent", header=None, engine=engine)

    # Rent calculator workbooks use IF formulas that cache 0 for unselected
    # utility options.  Extract the actual values from the formula text so
    # every option has a real allowance regardless of the saved selection.
    formula_values = _extract_if_formula_values(workbook_path)

    allowances = _parse_allowances(sheet, formula_values=formula_values)
    gross_rents = _parse_ami_rent_table(sheet)

    return RentSchedule(gross_rents=gross_rents, allowances=allowances)


def _parse_allowances(
    sheet: pd.DataFrame,
    formula_values: Dict[Tuple[int, int], float] | None = None,
) -> Dict[str, Dict[str, Dict[str, float]]]:
    allowances: Dict[str, Dict[str, Dict[str, float]]] = {}
    current_category = None

    # Detect layout variant (0-indexed rows):
    #   2025: category headers at row 14 ("Apartment Electricity only", etc.),
    #         option labels at rows 15-16, data at rows 17-22
    #   2024: option labels at row 14 ("Electric Stove", etc.),
    #         selection states at row 15, data at rows 16-21
    has_category_headers = False
    for ci in range(sheet.shape[1]):
        val = sheet.iloc[14, ci]
        if isinstance(val, str) and val.strip().lower() in HEADER_TO_CATEGORY:
            has_category_headers = True
            break

    first_data_row = 17 if has_category_headers else 16  # pandas 0-indexed

    for col_idx in range(sheet.shape[1]):
        header = sheet.iloc[14, col_idx]
        if isinstance(header, str):
            header_clean = header.strip()
            if header_clean:
                normalized = header_clean.lower()
                header_category = HEADER_TO_CATEGORY.get(normalized)
                if header_category:
                    current_category = header_category
                else:
                    option_category = LABEL_TO_CATEGORY.get(header_clean)
                    if option_category:
                        current_category = option_category

        # Option label: row 14 for 2024 (if recognized), rows 15-16 for 2025
        if not has_category_headers and isinstance(header, str) and header.strip() and LABEL_TO_CATEGORY.get(header.strip()):
            option = header  # 2024: row 15 IS the option label
        else:
            option = sheet.iloc[15, col_idx]
            if not isinstance(option, str) or not option.strip():
                fallback_option = sheet.iloc[16, col_idx] if sheet.shape[0] > 16 else None
                option = fallback_option if isinstance(fallback_option, str) else option
        if not isinstance(option, str):
            continue
        cleaned_option = option.strip()
        if not cleaned_option or cleaned_option.lower() == "select -->>":
            continue

        # The "Apartment Electricity only" column has no per-option label row (it's
        # the column's only paid option). When the dropdown in row 17 happens to
        # read "N/A or owner pays" (default state of a fresh HPD template), the
        # fallback above pulls that string in as if it were the option name —
        # which causes the column's allowance values to land under the wrong
        # category. Force the canonical "Tenant Pays" label when we recognise
        # the electricity category header but the option string came back as
        # the dropdown sentinel.
        if current_category == 'electricity' and cleaned_option == 'N/A or owner pays':
            cleaned_option = 'Tenant Pays'

        category_key = LABEL_TO_CATEGORY.get(cleaned_option) or current_category
        if not category_key:
            continue

        category_bucket = allowances.setdefault(category_key, {})
        bedroom_map = {}
        for offset, bedroom in enumerate(BEDROOM_LABELS):
            value = sheet.iloc[first_data_row + offset, col_idx]
            numeric = 0.0
            if not pd.isna(value):
                if isinstance(value, (int, float)):
                    numeric = float(value)
                elif isinstance(value, str):
                    value_str = value.strip()
                    if value_str.startswith('='):
                        match = re.search(r',\s*([-]?[0-9]+(?:\.[0-9]+)?)', value_str)
                        if match:
                            try:
                                numeric = float(match.group(1))
                            except ValueError:
                                numeric = 0.0
                    else:
                        try:
                            numeric = float(value_str)
                        except ValueError:
                            numeric = 0.0
            # Fallback: if cached formula result is 0, use the actual value
            # extracted from the IF formula text (see _extract_if_formula_values).
            if numeric == 0.0 and formula_values:
                extracted = formula_values.get((first_data_row + offset, col_idx))
                if extracted is not None and extracted > 0.0:
                    numeric = extracted

            bedroom_map[bedroom] = numeric
        category_bucket[cleaned_option] = bedroom_map
    return allowances


def _parse_ami_rent_table(sheet: pd.DataFrame) -> Dict[Tuple[float, str], float]:
    rents: Dict[Tuple[float, str], float] = {}
    current_ami = None
    for idx in range(sheet.shape[0]):
        cell = sheet.iloc[idx, 2]
        marker = sheet.iloc[idx, 3] if sheet.shape[1] > 3 else None

        if isinstance(cell, (int, float)) and isinstance(marker, str) and marker.strip().lower() == "of ami":
            current_ami = float(cell)
            continue

        if current_ami is None:
            continue

        if isinstance(cell, str):
            label = cell.strip()
            if label in BEDROOM_LABELS:
                gross_rent = sheet.iloc[idx, 6] if sheet.shape[1] > 6 else None
                if pd.isna(gross_rent):
                    raise ValueError(f"Missing gross rent for {current_ami*100:.0f}% AMI / {label}.")
                rents[(round(current_ami, 4), label)] = float(gross_rent)
    return rents


def _normalize_bedroom_label(bedrooms: float) -> str:
    try:
        bedrooms_int = int(round(float(bedrooms)))
    except (TypeError, ValueError):
        bedrooms_int = 0
    if bedrooms_int <= 0:
        return "studio"
    if bedrooms_int >= 5:
        return "5 BR"
    return f"{bedrooms_int} BR"


def _resolve_option_label(category: str, choice_key: str) -> str:
    mappings = {
        "electricity": ELECTRICITY_OPTIONS,
        "cooking": COOKING_OPTIONS,
        "heat": HEAT_OPTIONS,
        "hot_water": HOT_WATER_OPTIONS,
    }
    if category not in mappings:
        raise ValueError(f"Unknown utility category '{category}'.")
    options = mappings[category]
    if choice_key not in options:
        raise ValueError(f"Unsupported selection '{choice_key}' for category '{category}'.")
    return options[choice_key]


def compute_rents_for_assignments(
    schedule: RentSchedule,
    assignments: List[Dict[str, float]],
    utilities: Dict[str, str],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    updated_assignments: List[Dict[str, Any]] = []
    total_net = 0.0
    total_gross = 0.0
    total_allowance = 0.0
    category_totals: Dict[str, float] = {}
    category_labels: Dict[str, str] = {}

    for unit in assignments:
        ami = float(unit.get("assigned_ami"))
        bedrooms = unit.get("bedrooms", 0)
        components = schedule.rent_components(ami, bedrooms, utilities)
        gross = components['gross']
        net = components['net']
        allowance_total = components['allowance_total']
        allowance_details = []
        for category, detail in components['allowances'].items():
            amount = detail['amount']
            label = detail['label']
            allowance_details.append({
                'category': category,
                'label': label,
                'amount': round(amount, 2),
            })
            category_totals[category] = category_totals.get(category, 0.0) + amount
            category_labels[category] = label

        enriched = dict(unit)
        enriched['gross_rent'] = round(gross, 2)
        enriched['monthly_rent'] = round(net, 2)
        enriched['annual_rent'] = round(net * 12.0, 2)
        enriched['allowance_total'] = round(allowance_total, 2)
        enriched['allowances'] = allowance_details
        updated_assignments.append(enriched)

        total_net += net
        total_gross += gross
        total_allowance += allowance_total

    allowance_breakdown = {
        category: {
            'label': category_labels.get(category, ''),
            'monthly': round(amount, 2),
            'annual': round(amount * 12.0, 2),
        }
        for category, amount in category_totals.items()
    }

    totals = {
        'net_monthly': round(total_net, 2),
        'net_annual': round(total_net * 12.0, 2),
        'gross_monthly': round(total_gross, 2),
        'gross_annual': round(total_gross * 12.0, 2),
        'allowances_monthly': round(total_allowance, 2),
        'allowances_annual': round(total_allowance * 12.0, 2),
        'allowances_breakdown': allowance_breakdown,
    }

    return updated_assignments, totals


def save_rent_workbook_with_utilities(
    source_path: str,
    utilities: Dict[str, str],
    destination_path: str,
) -> str:
    """
    Persist utility selections into the rent workbook so Excel shows the same
    allowance choices the user picked in the UI.
    """
    if not source_path or not os.path.exists(source_path):
        raise FileNotFoundError(f"Rent workbook not found: {source_path}")

    source_ext = Path(source_path).suffix.lower()
    if source_ext == ".xlsb":
        raise ValueError("Writing utility selections into .xlsb workbooks is not supported.")

    workbook = load_excel_file(source_path)
    sheet = workbook["AMI & Rent"]

    mapping = [
        ("electricity", 17, 2, ELECTRICITY_OPTIONS),
        ("cooking", 17, 3, COOKING_OPTIONS),
        ("heat", 17, 5, HEAT_OPTIONS),
        ("hot_water", 17, 10, HOT_WATER_OPTIONS),
    ]

    for key, row, col, options in mapping:
        selection_key = utilities.get(key, "na")
        label = options.get(selection_key, options["na"])
        sheet.cell(row=row, column=col).value = label

    workbook.save(destination_path)
    return destination_path

